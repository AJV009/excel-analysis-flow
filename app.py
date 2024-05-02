from flask import Flask, request, render_template, jsonify, session, redirect, url_for, Response, stream_with_context
from flask_session import Session 
import pandas as pd
from lib.utils import is_sheet_small, encode_image, convert_to_pdf, page_number_mapping, get_img_from_pg_num, convert_to_csv, get_img_from_csv
from lib.util_agent import small_sheet_query_agent, search_term_extraction_agent, table_list, search_term_query_correction_agent, query_writer_agent, response_humanizer_agent, reset_usage, reset_logs
import json
import uuid
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
import hashlib
import glob
import sqlite3

load_dotenv()

app = Flask(__name__)
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
app.secret_key = os.urandom(24)  # Generate a random secret key
Session(app)

'''
TODO:
1. Error handling - Agents, Responses, Retries, etc.
2. Alternate paths for the agents
3. More than single "search" query, add tooling/examples for advanced search queries
4. Some issue with Docker based runtime
5. Direct streaming to the frontend. (Use openai stream in conjenction with Flask stream_with_context, generator functions and yield)
'''

@app.route('/login', methods=['GET'])
def login():
    return render_template('login.html')

# Route to handle login logic
@app.route('/login', methods=['POST'])
def verify_login():
    username = request.form['username']
    password = request.form['password']

    # Assuming you have an Excel file named 'credentials.xlsx' with 'Username' and 'Password' columns
    df = pd.read_excel('credentials.xlsx')

    # Verify credentials
    user = df[(df['username'] == username) & (df['password'] == password)]
    if not user.empty:
        session['username'] = username  # Store username in session
        return jsonify({'success': 'Logged in successfully'}), 200
    else:
        return jsonify({'error': 'Invalid username or password'}), 401

# Modify existing routes to check for login
@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))  # Redirect to login if not logged in
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():

    def generate_response():
        file = {}
        if 'file' not in request.files:
            yield json.dumps({'error': 'No file part'}).encode() + b'\n' 
            return
        file = request.files['file']
        if file.filename == '':
            yield json.dumps({'error': 'No selected file'}).encode() + b'\n'
            return
        if file:
            yield json.dumps({'success': 'Uploading the workbook. Please wait!'}).encode() + b'\n'
            # Calculate the file checksum
            file_content = file.read()
            file_checksum = hashlib.md5(file_content).hexdigest()
            file.seek(0)  # Reset file pointer to the beginning

            # Check if the file has already been processed
            existing_files = glob.glob('temp_files/**/*.xlsx', recursive=True)
            for existing_file in existing_files:
                with open(existing_file, 'rb') as f:
                    existing_file_content = f.read()
                    existing_file_checksum = hashlib.md5(existing_file_content).hexdigest()
                    if existing_file_checksum == file_checksum:
                        existing_cuuid = existing_file.split('/')[1]
                        # load the metadata file
                        metadata_file_path = f'temp_files/{existing_cuuid}/{existing_cuuid}_metadata.json'
                        if os.path.exists(metadata_file_path):    
                            file_metadata = {}
                            with open(metadata_file_path, 'r') as f:
                                file_metadata = json.load(f)
                            if 'cuuid' in file_metadata:
                                yield json.dumps({'success': 'indexed', 'cuuid': existing_cuuid}).encode() + b'\n'
                                return

            # load the workbook
            yield json.dumps({'success': 'Uploading the workbook. Please wait!'}).encode() + b'\n'
            print("Uploading the workbook. Please wait!")

            # create file paths and ids for tracking
            current_uuid = str(uuid.uuid4())

            # create a directory for the current file
            os.makedirs(f'temp_files/{current_uuid}', exist_ok=True)
            file_path = f'temp_files/{current_uuid}/{file.filename}'
            file.save(file_path)

            file_metadata = {
                "excel_file_path": file_path,
                "cuuid": "",
                "pdf_file_path": "",
                "small_sheets": {
                    "encoded_images_json": ""
                },
                "big_sheets": {
                    "csv_file_meta": [],
                    "sqllite_db_path": ""
                }
            }

            # load to openpyxl for smaller sheets filtering
            workbook = load_workbook(file_path)
            yield json.dumps({'success': 'Indexing: Workbook uploaded, now processing sheets. Please wait!'}).encode() + b'\n'
            print("Indexing: Workbook uploaded, now processing sheets. Please wait!")

            # complete workbook to pdf conversion
            file_metadata['pdf_file_path'] = convert_to_pdf(file_path, current_uuid)

            # workbook page mapping
            page_mapping = page_number_mapping(workbook.sheetnames)

            def process_small_sheets():
                small_sheet_images_paths = []
                for pg_num in [page_mapping[sheet_name] for sheet_name in [sheet.title for sheet in workbook if is_sheet_small(sheet)]]:
                    small_sheet_images_paths.append(get_img_from_pg_num(file_metadata['pdf_file_path'], pg_num, page_mapping, current_uuid))

                # small sheet images encoder
                small_sheet_images_context_array = []
                for img in small_sheet_images_paths:
                    if img is not None:
                        image_base64 = encode_image(img)
                        image_encoding = f"data:image/jpeg;base64,{image_base64}"
                        small_sheet_images_context_array.append({
                            "image_path": img,
                            "image_encoding": image_encoding,
                        })
                # save the images context array
                file_metadata['small_sheets']['encoded_images_json'] = f'temp_files/{current_uuid}/small_sheet_images_context_array.json'
                with open(file_metadata['small_sheets']['encoded_images_json'], 'w') as f:
                    json.dump(small_sheet_images_context_array, f)

            def process_big_sheets():
                # csv generation & sample image extraction
                for sheet in workbook:
                    if is_sheet_small(sheet, return_big_sheets=True):
                        csv_file = f'temp_files/{current_uuid}/{sheet.title}.csv'
                        file_metadata['big_sheets']['csv_file_meta'].append({
                            "sheet_name": sheet.title,
                            "csv_file_path": convert_to_csv(sheet, csv_file),
                            "csv_sample_image": get_img_from_csv(csv_file, page_mapping[sheet.title], page_mapping, current_uuid)
                        })

                # store in sqllite db
                file_metadata['big_sheets']['sqllite_db_path'] = f'temp_files/{current_uuid}/workbook.db'
                conn = sqlite3.connect(file_metadata['big_sheets']['sqllite_db_path'])
                for csv_meta in file_metadata['big_sheets']['csv_file_meta']:
                    df = pd.read_csv(csv_meta['csv_file_path'])
                    df.to_sql(csv_meta['sheet_name'], conn, if_exists='replace', index=False, method='multi')
                conn.close()

            # Process small sheets if they exist
            if any(is_sheet_small(sheet) for sheet in workbook):
                yield json.dumps({'success': 'Indexing: Processing small sheets. Please wait!'}).encode() + b'\n'
                print("Indexing: Processing small sheets. Please wait!")
                process_small_sheets()

            # Process big sheets if they exist
            if any(is_sheet_small(sheet, return_big_sheets=True) for sheet in workbook):
                yield json.dumps({'success': 'Indexing: Processing big sheets. Please wait!'}).encode() + b'\n'
                print("Indexing: Processing big sheets. Please wait!")
                process_big_sheets()

            yield json.dumps({'success': f'Indexing: Processed and encoded all sheets. Saving metadata for user {current_uuid}. Please wait!'}).encode() + b'\n'
            print(f"Indexing: Processed and encoded all sheets. Saving metadata for user {current_uuid}. Please wait!")

            # Adding cuuid to the metadata in the end to indicate successful indexing
            file_metadata['cuuid'] = current_uuid
            # save file_metadata to json
            with open(f'temp_files/{current_uuid}/{current_uuid}_metadata.json', 'w') as f:
                json.dump(file_metadata, f)

            yield json.dumps({'success': 'indexed', 'cuuid': current_uuid}).encode() + b'\n'
            print("File indexed")
            return

    return Response(stream_with_context(generate_response()), content_type='application/json')

@app.route('/ask', methods=['POST'])
def ask_question():
    
    def generate_response():
        question = request.form['question']
        cuuid = request.form['cuuid']

        metadata_file_path = f'temp_files/{cuuid}/{cuuid}_metadata.json'
        
        if os.path.exists(metadata_file_path):
            with open(metadata_file_path, 'r') as f:
                file_metadata = json.load(f)
                if 'cuuid' not in file_metadata:
                    yield json.dumps({'error': 'File not indexed yet'}).encode() + b'\n'
                    return

        reset_usage(cuuid)
        reset_logs(cuuid)
        yield json.dumps({'success': 'Analyzing: Processing the query. Please wait!', 'action': "processing"}).encode() + b'\n'

        # load metadata
        metadata = {}
        with open(f'temp_files/{cuuid}/{cuuid}_metadata.json', 'r') as f:
            metadata = json.load(f)
        
        yield json.dumps({'success': 'Analyzing: Checking against small sheets. Please wait!', 'action': "processing"}).encode() + b'\n'
        # check against small sheets first, if they exist
        small_sheet_images = {}
        if len(metadata['small_sheets']['encoded_images_json']) > 0:
            # check if metadata['small_sheets']['encoded_images_json'] exists
            if os.path.exists(metadata['small_sheets']['encoded_images_json']):
                with open(metadata['small_sheets']['encoded_images_json'], 'r') as f:
                    small_sheet_images = json.load(f)

                # check if the question can be answered from small sheets
                message_image_array = []
                for img in small_sheet_images:
                    message_image_array.append({
                        "type": "image_url",
                        "image_url": {
                            "url": img['image_encoding']
                        }
                    })
                
                # call small sheet agent
                small_sheet_response = small_sheet_query_agent(question, message_image_array, cuuid)
                if small_sheet_response != 'no_answer_found':
                    yield from response_humanizer_agent(question, small_sheet_response, cuuid)
                    return
        
        # check against big sheets, if they exist
        yield json.dumps({'success': 'Analyzing: No related data in small sheets, checking against large sheets. Please wait!', 'action': "processing"}).encode() + b'\n'
        if len(metadata['big_sheets']['sqllite_db_path']) > 0:
            # fix search query flow here
            
            yield json.dumps({'success': 'Analyzing: Correcting query. Please wait!', 'action': "processing"}).encode() + b'\n'
            search_term = search_term_extraction_agent(query=question, cuuid=cuuid)
            question = search_term_query_correction_agent(query=question, search_term=search_term, metadata=metadata, cuuid=cuuid)
            
            db = sqlite3.connect(metadata['big_sheets']['sqllite_db_path'])
            cursor = db.cursor()
            search_term_table_names = table_list(metadata=metadata, search_term=search_term)
            table_detail_mapping = {}
            for table in search_term_table_names:
                for file_meta in metadata['big_sheets']['csv_file_meta']:
                    if table in file_meta['sheet_name']:
                        tab_q = f'PRAGMA table_info("{table}")'
                        cursor.execute(tab_q)
                        table_description = cursor.fetchall()
                        for col in table_description:
                            table_detail_mapping[table] = {}
                            table_detail_mapping[table]["column_details"] = []
                            table_detail_mapping[table]["column_details"].append({ 'column_name': col[1], 'column_type': col[2] })
                        table_detail_mapping[table]["table_name"] = table
                        table_detail_mapping[table]['sample_image'] = file_meta['csv_sample_image']
                        break
            yield json.dumps({'success': 'Analyzing: Checking against individual large sheets. Please wait!', 'action': "processing"}).encode() + b'\n'
            multiple_sql_queries = query_writer_agent(question, table_detail_mapping, cuuid)
            
            # execute all queries, use try except block, put all results in a list
            results = []
            for query in multiple_sql_queries:
                try:
                    print(query["query"])
                    cursor.execute(query["query"])
                    result = cursor.fetchall()
                    print(result)
                    results.append({
                        "table_name": query["table_name"],
                        "query": query["query"],
                        "result": result
                    })
                except Exception as e:
                    results.append({
                        "table_name": query["table_name"],
                        "query": query["query"],
                        "result": str(e)
                    })
            db.close()
            yield from response_humanizer_agent(question, results, cuuid)
            return

    return Response(stream_with_context(generate_response()), content_type='application/json')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002)
